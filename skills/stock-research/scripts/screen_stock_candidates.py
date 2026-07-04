#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import requests


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

# Screening universe snapshots are locked to the single Eastmoney webguest
# clist provider defined in screen_stock_common. Financial enrich can use
# market-specific Eastmoney statement APIs; market snapshots cannot mix sources.
from fetch_risk_free_rate import fetch_risk_free_rate  # noqa: E402
import screen_stock_providers  # noqa: E402
from screen_stock_common import (  # noqa: E402
    CURRENCY_ALIASES,
    DEFAULT_FINANCIAL_CACHE_MAX_AGE_DAYS,
    DEFAULT_MIN_MARKET_CAP,
    DEFAULT_PAGE_SIZE,
    DEFAULT_PROGRESS_INTERVAL,
    DEFAULT_REQUEST_DELAY_SECONDS,
    EASTMONEY_LIST_FIELDS,
    EASTMONEY_QUOTE_UT,
    EASTMONEY_WBP2U,
    ENRICH_FILENAME,
    INDUSTRY_REVIEW_KEYWORDS,
    LOGGER,
    MARKET_CONFIGS,
    MARKET_SNAPSHOT_ENDPOINT,
    MARKET_SNAPSHOT_PROVIDER,
    MIN_FINANCIAL_YEARS,
    SORT_FIELDS,
    UNIVERSE_FILENAME,
    MarketSnapshot,
    append_jsonl,
    checkpoint_status_counts,
    configure_progress_logger,
    configure_session,
    default_cache_dir,
    default_run_date,
    default_run_dir,
    eastmoney_price,
    eastmoney_ratio,
    is_industry_review_required,
    json_safe,
    log_screen_progress,
    mean,
    normalize_currency,
    now_iso,
    parse_json_or_jsonp,
    read_jsonl,
    round_or_none,
    safe_ratio,
    snapshot_from_dict,
    snapshot_to_dict,
    sleep_if_needed,
    stdev,
    ticker_for_market,
    to_float,
    to_scaled_float,
    write_json_path,
)
from screen_stock_metrics import (  # noqa: E402
    build_result_item,
    classify_candidate,
    complete_financial_rows,
    compute_metrics,
    failed_market_snapshots,
    financial_currency,
    financial_fetch_blockers,
    financial_fetch_failed_item,
    quick_reject_reasons,
    refresh_result_item_contract,
    sort_results,
)
from screen_stock_parsers import (  # noqa: E402
    parse_market_snapshot,
)
from screen_stock_providers import (  # noqa: E402
    EASTMONEY_DATA_URL,
    EASTMONEY_DATA_V1_URL,
    EASTMONEY_HK_DATA_URL,
    EASTMONEY_WEBGUEST_LIST_URL,
    annotate_market_cap_universe,
    fetch_a_company_type,
    fetch_a_financial_rows,
    fetch_a_statement_rows,
    fetch_financial_rows,
    fetch_hk_company_type,
    fetch_hk_financial_rows,
    fetch_market_snapshots,
    fetch_us_financial_rows,
    get_json,
    get_with_retry,
)


def financial_cache_path(cache_dir: Path, snapshot: MarketSnapshot) -> Path:
    return cache_dir / snapshot.market / f"{snapshot.code}.json"


def cache_payload_is_fresh(payload: dict[str, Any], max_age_days: float | None) -> bool:
    if max_age_days is None or max_age_days <= 0:
        return True
    fetched_at = payload.get("fetched_at")
    if not isinstance(fetched_at, str) or not fetched_at.strip():
        return False
    try:
        parsed = datetime.fromisoformat(fetched_at.replace("Z", "+00:00"))
    except ValueError:
        return False
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return datetime.now(timezone.utc) - parsed <= timedelta(days=max_age_days)


def load_financial_rows_from_cache(
    cache_dir: Path | None,
    snapshot: MarketSnapshot,
    *,
    max_age_days: float | None = DEFAULT_FINANCIAL_CACHE_MAX_AGE_DAYS,
) -> list[dict[str, Any]] | None:
    if cache_dir is None:
        return None
    path = financial_cache_path(cache_dir, snapshot)
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not cache_payload_is_fresh(payload, max_age_days):
        return None
    rows = payload.get("financial_rows")
    if not isinstance(rows, list):
        return None
    return rows


def write_financial_rows_cache(cache_dir: Path | None, snapshot: MarketSnapshot, rows: list[dict[str, Any]]) -> None:
    if cache_dir is None:
        return
    payload = {
        "ticker": snapshot.ticker,
        "code": snapshot.code,
        "market": snapshot.market,
        "fetched_at": now_iso(),
        "source": "eastmoney",
        "report_periods": [row.get("year") for row in rows if row.get("year")],
        "financial_rows": rows,
    }
    write_json_path(financial_cache_path(cache_dir, snapshot), payload)


def fetch_financial_rows_cached(
    snapshot: MarketSnapshot,
    *,
    cache_dir: Path | None,
    cache_max_age_days: float | None = DEFAULT_FINANCIAL_CACHE_MAX_AGE_DAYS,
    session: requests.sessions.Session,
    request_delay: float,
    timeout: float,
) -> tuple[list[dict[str, Any]], bool]:
    cached = load_financial_rows_from_cache(cache_dir, snapshot, max_age_days=cache_max_age_days)
    if cached is not None:
        return cached, True
    rows = fetch_financial_rows(snapshot, session=session, request_delay=request_delay, timeout=timeout)
    write_financial_rows_cache(cache_dir, snapshot, rows)
    return rows, False


def write_universe(
    run_dir: Path,
    *,
    run_date: str,
    markets: list[str],
    snapshots: list[MarketSnapshot],
    errors: list[dict[str, Any]],
    min_market_cap: float,
) -> None:
    write_json_path(
        run_dir / UNIVERSE_FILENAME,
        {
            "run_date": run_date,
            "fetched_at": now_iso(),
            "markets": markets,
            "market_snapshot_provider": MARKET_SNAPSHOT_PROVIDER,
            "market_snapshot_endpoint": MARKET_SNAPSHOT_ENDPOINT,
            "min_market_cap": min_market_cap,
            "snapshot_count": len(snapshots),
            "errors": errors,
            "snapshots": [snapshot_to_dict(snapshot) for snapshot in snapshots],
        },
    )


def reset_run_checkpoint(run_dir: Path | None) -> None:
    if run_dir is None:
        return
    for filename in (UNIVERSE_FILENAME, ENRICH_FILENAME, "progress.json"):
        path = run_dir / filename
        if path.exists():
            path.unlink()


def load_universe(run_dir: Path) -> tuple[list[MarketSnapshot], list[dict[str, Any]], dict[str, Any]]:
    path = run_dir / UNIVERSE_FILENAME
    payload = json.loads(path.read_text(encoding="utf-8"))
    snapshots = [snapshot_from_dict(item) for item in payload.get("snapshots", [])]
    errors = payload.get("errors") or []
    return snapshots, errors, payload


def latest_checkpoint_rows(run_dir: Path) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for row in read_jsonl(run_dir / ENRICH_FILENAME):
        ticker = row.get("ticker")
        if ticker:
            latest[str(ticker)] = row
    return latest


def checkpoint_result(
    run_dir: Path | None,
    *,
    run_date: str,
    snapshot: MarketSnapshot,
    enrich_status: str,
    result: dict[str, Any],
    attempt: int,
    error: str | None = None,
    from_cache: bool = False,
) -> None:
    if run_dir is None:
        return
    row = {
        "run_date": run_date,
        "fetched_at": now_iso(),
        "ticker": snapshot.ticker,
        "market": snapshot.market,
        "code": snapshot.code,
        "enrich_status": enrich_status,
        "attempt": attempt,
        "from_cache": from_cache,
        "result": result,
    }
    if error:
        row["error"] = error
    append_jsonl(run_dir / ENRICH_FILENAME, row)


def screen_stocks(
    *,
    markets: list[str],
    max_count: int,
    sort_by: str,
    min_market_cap: float,
    request_delay: float = 0.0,
    session: requests.sessions.Session | None = None,
    timeout: float = 20,
    run_date: str | None = None,
    run_dir: Path | None = None,
    resume: bool = False,
    cache_dir: Path | None = None,
    cache_max_age_days: float | None = DEFAULT_FINANCIAL_CACHE_MAX_AGE_DAYS,
    progress_interval: int = 0,
    logger: logging.Logger | None = None,
) -> dict[str, Any]:
    run_date = run_date or default_run_date()
    normalized_markets = [market.upper() for market in markets]
    unsupported = [market for market in normalized_markets if market not in MARKET_CONFIGS]
    supported_markets = [market for market in normalized_markets if market in MARKET_CONFIGS]
    client = configure_session(session or requests.Session())
    if run_dir is not None and not resume:
        reset_run_checkpoint(run_dir)

    if resume and run_dir is not None and (run_dir / UNIVERSE_FILENAME).exists():
        snapshots, errors, universe_payload = load_universe(run_dir)
        universe_source = "checkpoint"
        universe_fetched_at = universe_payload.get("fetched_at")
        universe_run_date = universe_payload.get("run_date")
        universe_provider = universe_payload.get("market_snapshot_provider")
        universe_endpoint = universe_payload.get("market_snapshot_endpoint")
        if universe_run_date and universe_run_date != run_date:
            return {
                "ok": False,
                "fetched_at": now_iso(),
                "markets": supported_markets,
                "unsupported_markets": unsupported,
                "source": "secondary_market_data",
                "data_limits": {
                    "secondary_source_only": True,
                    "max_count": max_count,
                    "run_date": run_date,
                    "run_dir": str(run_dir),
                    "cache_dir": str(cache_dir) if cache_dir is not None else None,
                    "financial_cache_max_age_days": cache_max_age_days,
                    "resume": resume,
                    "universe_source": universe_source,
                    "universe_fetched_at": universe_fetched_at,
                    "universe_run_date": universe_run_date,
                    "market_snapshot_provider": MARKET_SNAPSHOT_PROVIDER,
                    "market_snapshot_endpoint": MARKET_SNAPSHOT_ENDPOINT,
                    "checkpoint_market_snapshot_provider": universe_provider,
                    "checkpoint_market_snapshot_endpoint": universe_endpoint,
                    "snapshot_sources": sorted({snapshot.source for snapshot in snapshots if snapshot.source}),
                    "financial_analysis_scope": "all_rows_at_or_above_market_cap_floor",
                    "financial_analysis_selected_count": 0,
                    "analyzed_count": 0,
                    "financial_fetch_attempts": 0,
                    "min_market_cap": min_market_cap,
                    "request_delay_seconds": request_delay,
                    "progress_interval": progress_interval,
                    "errors": errors,
                },
                "risk_free_rates": {},
                "summary": {
                    "total": 0,
                    "candidate": 0,
                    "rejected": 0,
                    "insufficient_data": 0,
                    "industry_review_required": 0,
                },
                "phase": "resume_checkpoint",
                "error": (
                    f"Cannot resume run_date {run_date} from checkpoint run_date {universe_run_date}. "
                    "Start a new date directory or pass the checkpoint run_date explicitly."
                ),
                "results": [],
            }
        if universe_provider != MARKET_SNAPSHOT_PROVIDER or universe_endpoint != MARKET_SNAPSHOT_ENDPOINT:
            return {
                "ok": False,
                "fetched_at": now_iso(),
                "markets": supported_markets,
                "unsupported_markets": unsupported,
                "source": "secondary_market_data",
                "data_limits": {
                    "secondary_source_only": True,
                    "max_count": max_count,
                    "run_date": run_date,
                    "run_dir": str(run_dir),
                    "cache_dir": str(cache_dir) if cache_dir is not None else None,
                    "financial_cache_max_age_days": cache_max_age_days,
                    "resume": resume,
                    "universe_source": universe_source,
                    "universe_fetched_at": universe_fetched_at,
                    "universe_run_date": universe_run_date,
                    "market_snapshot_provider": MARKET_SNAPSHOT_PROVIDER,
                    "market_snapshot_endpoint": MARKET_SNAPSHOT_ENDPOINT,
                    "checkpoint_market_snapshot_provider": universe_provider,
                    "checkpoint_market_snapshot_endpoint": universe_endpoint,
                    "snapshot_sources": sorted({snapshot.source for snapshot in snapshots if snapshot.source}),
                    "financial_analysis_scope": "all_rows_at_or_above_market_cap_floor",
                    "financial_analysis_selected_count": 0,
                    "analyzed_count": 0,
                    "financial_fetch_attempts": 0,
                    "min_market_cap": min_market_cap,
                    "request_delay_seconds": request_delay,
                    "progress_interval": progress_interval,
                    "errors": errors,
                },
                "risk_free_rates": {},
                "summary": {
                    "total": 0,
                    "candidate": 0,
                    "rejected": 0,
                    "insufficient_data": 0,
                    "industry_review_required": 0,
                },
                "phase": "resume_checkpoint",
                "error": (
                    "Cannot resume checkpoint from a different market snapshot provider. "
                    f"Expected {MARKET_SNAPSHOT_PROVIDER} at {MARKET_SNAPSHOT_ENDPOINT}."
                ),
                "results": [],
            }
    else:
        snapshots, errors = fetch_market_snapshots(
            markets=supported_markets,
            max_count=max_count,
            sort_by="market_cap",
            min_market_cap=min_market_cap,
            request_delay=request_delay,
            session=client,
            timeout=timeout,
        )
        annotate_market_cap_universe(snapshots)
        universe_source = "live"
        universe_fetched_at = now_iso()
        if run_dir is not None:
            write_universe(
                run_dir,
                run_date=run_date,
                markets=supported_markets,
                snapshots=snapshots,
                errors=errors,
                min_market_cap=min_market_cap,
            )
    for market in unsupported:
        errors.append({"market": market, "phase": "market_snapshot", "error": "unsupported market"})

    if logger is not None:
        logger.info(
            "[stock-screen] universe %s total=%s markets=%s run_date=%s",
            universe_source,
            len(snapshots),
            ",".join(supported_markets) or "-",
            run_date,
        )

    rates: dict[str, dict[str, Any]] = {}
    results_by_ticker: dict[str, dict[str, Any]] = {}
    checkpoint_rows = latest_checkpoint_rows(run_dir) if run_dir is not None else {}
    if resume:
        for ticker, row in checkpoint_rows.items():
            result = row.get("result")
            if isinstance(result, dict):
                results_by_ticker[ticker] = refresh_result_item_contract(result, min_market_cap=min_market_cap)
    financial_fetch_attempts = 0
    for processed_count, snapshot in enumerate(snapshots, start=1):
        previous_row = checkpoint_rows.get(snapshot.ticker)
        if resume and previous_row and previous_row.get("enrich_status") == "ok":
            log_screen_progress(
                logger,
                event="skipped",
                processed=processed_count,
                total=len(snapshots),
                ticker=snapshot.ticker,
                checkpoint_rows=checkpoint_rows,
                financial_fetch_attempts=financial_fetch_attempts,
                interval=progress_interval,
            )
            continue

        fetch_blockers = financial_fetch_blockers(snapshot, min_market_cap)
        should_analyze = not fetch_blockers and snapshot.selected_for_financial_analysis
        financial_rows: list[dict[str, Any]] = []
        from_cache = False
        attempt = int((previous_row or {}).get("attempt") or 0) + 1
        if should_analyze:
            financial_fetch_attempts += 1
            try:
                financial_rows, from_cache = fetch_financial_rows_cached(
                    snapshot,
                    cache_dir=cache_dir,
                    cache_max_age_days=cache_max_age_days,
                    session=client,
                    request_delay=request_delay,
                    timeout=timeout,
                )
            except Exception as exc:
                result = financial_fetch_failed_item(snapshot, min_market_cap=min_market_cap, error=str(exc))
                results_by_ticker[snapshot.ticker] = result
                checkpoint_result(
                    run_dir,
                    run_date=run_date,
                    snapshot=snapshot,
                    enrich_status="failed",
                    result=result,
                    attempt=attempt,
                    error=str(exc),
                )
                checkpoint_rows[snapshot.ticker] = {
                    "ticker": snapshot.ticker,
                    "enrich_status": "failed",
                    "attempt": attempt,
                    "result": result,
                }
                log_screen_progress(
                    logger,
                    event="failed",
                    processed=processed_count,
                    total=len(snapshots),
                    ticker=snapshot.ticker,
                    checkpoint_rows=checkpoint_rows,
                    financial_fetch_attempts=financial_fetch_attempts,
                    interval=progress_interval,
                    force=True,
                )
                continue

        fin_currency = financial_currency(financial_rows, snapshot.currency)
        if fin_currency not in rates:
            rates[fin_currency] = fetch_risk_free_rate(fin_currency, timeout=timeout)
        result = build_result_item(
            snapshot,
            financial_rows=financial_rows,
            risk_free=rates.get(fin_currency),
            min_market_cap=min_market_cap,
        )
        results_by_ticker[snapshot.ticker] = result
        checkpoint_result(
            run_dir,
            run_date=run_date,
            snapshot=snapshot,
            enrich_status="ok",
            result=result,
            attempt=attempt,
            from_cache=from_cache,
        )
        checkpoint_rows[snapshot.ticker] = {
            "ticker": snapshot.ticker,
            "enrich_status": "ok",
            "attempt": attempt,
            "result": result,
        }
        log_screen_progress(
            logger,
            event="ok_from_cache" if from_cache else "ok",
            processed=processed_count,
            total=len(snapshots),
            ticker=snapshot.ticker,
            checkpoint_rows=checkpoint_rows,
            financial_fetch_attempts=financial_fetch_attempts,
            interval=progress_interval,
        )

    results = [results_by_ticker[snapshot.ticker] for snapshot in snapshots if snapshot.ticker in results_by_ticker]
    results = sort_results(results, sort_by)
    snapshot_sources = sorted({snapshot.source for snapshot in snapshots if snapshot.source})
    failed_markets = failed_market_snapshots(supported_markets, errors)
    no_supported_markets = bool(normalized_markets) and not supported_markets
    all_supported_markets_failed = bool(supported_markets) and not snapshots and set(supported_markets) <= failed_markets
    no_snapshots_with_errors = bool(supported_markets) and not snapshots and bool(errors)
    ok = not (no_supported_markets or all_supported_markets_failed or no_snapshots_with_errors)
    payload = {
        "ok": ok,
        "fetched_at": now_iso(),
        "markets": supported_markets,
        "unsupported_markets": unsupported,
        "source": "secondary_market_data",
        "data_limits": {
            "secondary_source_only": True,
            "max_count": max_count,
            "run_date": run_date,
            "run_dir": str(run_dir) if run_dir is not None else None,
            "cache_dir": str(cache_dir) if cache_dir is not None else None,
            "financial_cache_max_age_days": cache_max_age_days,
            "resume": resume,
            "universe_source": universe_source,
            "universe_fetched_at": universe_fetched_at,
            "market_snapshot_provider": MARKET_SNAPSHOT_PROVIDER,
            "market_snapshot_endpoint": MARKET_SNAPSHOT_ENDPOINT,
            "snapshot_sources": snapshot_sources,
            "financial_analysis_scope": "all_rows_at_or_above_market_cap_floor",
            "financial_analysis_selected_count": sum(1 for item in snapshots if item.selected_for_financial_analysis),
            "analyzed_count": sum(1 for item in results if item.get("analyzed")),
            "financial_fetch_attempts": financial_fetch_attempts,
            "min_market_cap": min_market_cap,
            "request_delay_seconds": request_delay,
            "progress_interval": progress_interval,
            "errors": errors,
        },
        "risk_free_rates": rates,
        "summary": {
            "total": len(results),
            "candidate": sum(1 for item in results if item["status"] == "CANDIDATE"),
            "rejected": sum(1 for item in results if item["status"] == "REJECTED"),
            "insufficient_data": sum(1 for item in results if item["status"] == "INSUFFICIENT_DATA"),
            "industry_review_required": sum(1 for item in results if item["status"] == "INDUSTRY_REVIEW_REQUIRED"),
        },
        "results": results,
    }
    if no_supported_markets:
        payload["phase"] = "market_snapshot"
        payload["error"] = "No supported markets requested."
    elif all_supported_markets_failed:
        payload["phase"] = "market_snapshot"
        payload["error"] = "No market snapshots fetched from Eastmoney for requested markets."
    elif no_snapshots_with_errors:
        payload["phase"] = "market_snapshot"
        payload["error"] = "No market snapshots fetched; see data_limits.errors."
    if logger is not None:
        ok_count, failed_count = checkpoint_status_counts(checkpoint_rows)
        logger.info(
            "[stock-screen] complete total=%s ok=%s failed=%s results=%s candidates=%s run_dir=%s",
            len(snapshots),
            ok_count,
            failed_count,
            len(results),
            payload["summary"]["candidate"],
            str(run_dir) if run_dir is not None else "-",
        )
    return payload


def write_json(path: str, payload: dict[str, Any]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_xlsx(path: str, payload: dict[str, Any]) -> None:
    from openpyxl import Workbook

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    base_columns = [
        "ticker",
        "name",
        "market",
        "currency",
        "financial_currency",
        "status",
        "price",
        "market_cap",
        "market_cap_rank",
        "market_cap_percentile",
        "analyzed",
        "industry",
        "reject_reasons",
        "data_gaps",
        "error_phase",
        "error",
        "source",
    ]
    metric_columns = [
        "pe",
        "pb",
        "roe_mean",
        "roe_std",
        "roe_stability",
        "roic_mean",
        "avg_net_profit_5y",
        "avg_ocf_5y",
        "avg_fcf_5y",
        "ocf_to_profit",
        "fcf_to_profit",
        "market_cap_to_avg_profit",
        "market_cap_to_avg_fcf",
        "expected_return",
        "debt_to_assets",
        "goodwill_to_equity",
        "risk_free_multiple_cap",
        "complete_financial_years",
        "negative_fcf_years",
    ]

    def row_for(item: dict[str, Any]) -> list[Any]:
        row: list[Any] = []
        metrics = item.get("metrics") or {}
        for column in base_columns:
            value = item.get(column)
            if isinstance(value, list):
                value = ", ".join(str(part) for part in value)
            row.append(value)
        for column in metric_columns:
            row.append(metrics.get(column))
        return row

    columns = base_columns + metric_columns

    all_sheet = workbook.create_sheet("All")
    all_sheet.append(columns)
    for item in payload["results"]:
        all_sheet.append(row_for(item))

    for market in payload.get("markets", []):
        sheet = workbook.create_sheet(market[:31])
        sheet.append(columns)
        for item in payload["results"]:
            if item.get("market") == market:
                sheet.append(row_for(item))

    for status in ("CANDIDATE", "REJECTED", "INSUFFICIENT_DATA", "INDUSTRY_REVIEW_REQUIRED"):
        sheet = workbook.create_sheet(status.title().replace("_", ""))
        sheet.append(columns)
        for item in payload["results"]:
            if item["status"] != status:
                continue
            sheet.append(row_for(item))

    raw_sheet = workbook.create_sheet("RawSnapshots")
    raw_sheet.append(["ticker", "metrics_json"])
    for item in payload["results"]:
        raw_sheet.append([item["ticker"], json.dumps(json_safe(item.get("metrics", {})), ensure_ascii=False)])
    workbook.save(output_path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Screen stock candidates for quality value research.")
    parser.add_argument("--markets", nargs="+", default=["SH", "SZ", "HK"], help="Markets to screen. Default: SH SZ HK.")
    parser.add_argument(
        "--max-count",
        type=int,
        default=300,
        help="Maximum market snapshots to keep per market after sorting. Use 0 to page until --min-market-cap is crossed.",
    )
    parser.add_argument("--sort-by", choices=["market_cap", "expected_return"], default="market_cap")
    parser.add_argument("--min-market-cap", type=float, default=DEFAULT_MIN_MARKET_CAP)
    parser.add_argument("--request-delay", type=float, default=DEFAULT_REQUEST_DELAY_SECONDS)
    parser.add_argument(
        "--progress-interval",
        type=int,
        default=DEFAULT_PROGRESS_INTERVAL,
        help="Print progress to stderr every N processed rows. Use 0 to disable periodic progress.",
    )
    parser.add_argument("--timeout", type=float, default=20)
    parser.add_argument("--run-date", default=default_run_date(), help="Run date partition in YYYY-MM-DD format.")
    parser.add_argument(
        "--run-dir",
        help="Checkpoint run directory. Defaults to reports/stock-screen/<run-date> when --resume is used.",
    )
    parser.add_argument("--resume", action="store_true", help="Resume from the run directory checkpoint files.")
    parser.add_argument(
        "--cache-dir",
        help="Financial-row cache directory. Defaults to reports/stock-screen/cache/financials when checkpointing.",
    )
    parser.add_argument(
        "--cache-max-age-days",
        type=float,
        default=DEFAULT_FINANCIAL_CACHE_MAX_AGE_DAYS,
        help="Maximum age for financial-row cache reuse. Default: 7 days. Use 0 or less to disable TTL.",
    )
    parser.add_argument("--json-out")
    parser.add_argument("--xlsx-out")
    args = parser.parse_args(argv)

    run_dir = Path(args.run_dir) if args.run_dir else (default_run_dir(args.run_date) if args.resume else None)
    cache_dir = Path(args.cache_dir) if args.cache_dir else (default_cache_dir() if run_dir is not None else None)
    logger = configure_progress_logger()
    progress_interval = max(args.progress_interval, 0)

    payload = screen_stocks(
        markets=args.markets,
        max_count=args.max_count,
        sort_by=args.sort_by,
        min_market_cap=args.min_market_cap,
        request_delay=args.request_delay,
        timeout=args.timeout,
        run_date=args.run_date,
        run_dir=run_dir,
        resume=args.resume,
        cache_dir=cache_dir,
        cache_max_age_days=args.cache_max_age_days,
        progress_interval=progress_interval,
        logger=logger,
    )
    if args.json_out:
        write_json(args.json_out, payload)
    xlsx_out = args.xlsx_out
    if xlsx_out is None and run_dir is not None:
        xlsx_out = str(run_dir / f"stock-screen-{args.run_date}.xlsx")
    if xlsx_out:
        write_xlsx(xlsx_out, payload)

    stdout_payload = {key: value for key, value in payload.items() if key != "results"}
    stdout_payload["results"] = payload["results"][: min(50, len(payload["results"]))]
    print(json.dumps(json_safe(stdout_payload), ensure_ascii=False, indent=2))
    return 0 if payload.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
