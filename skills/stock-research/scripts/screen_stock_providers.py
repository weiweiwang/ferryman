from __future__ import annotations

import math
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import requests

from screen_stock_common import (
    A_SHARE_SELECTOR_FIELDS,
    DEFAULT_HEADERS,
    DEFAULT_PAGE_SIZE,
    EASTMONEY_LIST_FIELDS,
    EASTMONEY_QUOTE_UT,
    EASTMONEY_WBP2U,
    MARKET_CONFIGS,
    MIN_FINANCIAL_YEARS,
    SORT_FIELDS,
    ListingEntry,
    MarketSnapshot,
    configure_session,
    normalize_currency,
    parse_json_or_jsonp,
    sleep_if_needed,
    to_float,
)
from screen_stock_parsers import (
    parse_a_share_selector_snapshot,
    parse_market_snapshot,
    parse_tencent_hk_snapshot,
    parse_tencent_quote_records,
)


EASTMONEY_LIST_URL = "https://push2.eastmoney.com/api/qt/clist/get"
EASTMONEY_DATA_URL = "https://datacenter.eastmoney.com/securities/api/data/get"
EASTMONEY_HK_DATA_URL = "https://emh5.eastmoney.com/api/GangGu/CaiWu"
EASTMONEY_XUANGU_LIST_URL = "https://data.eastmoney.com/dataapi/xuangu/list"
TENCENT_QUOTE_URL = "https://qt.gtimg.cn/q="
HKEX_SECURITIES_LIST_URL = "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx"


def get_with_retry(
    session: requests.sessions.Session,
    url: str,
    *,
    params: dict[str, Any] | None = None,
    headers: dict[str, str] | None = None,
    timeout: float = 30,
    attempts: int = 3,
    request_delay: float = 0.0,
) -> requests.Response:
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            response = session.get(url, params=params, headers=headers, timeout=timeout)
            response.raise_for_status()
            return response
        except Exception as exc:
            last_error = exc
            if attempt < attempts:
                sleep_if_needed(max(request_delay, 0.5 * attempt))
        finally:
            sleep_if_needed(request_delay)
    assert last_error is not None
    raise last_error


def fetch_hkex_listings(
    session: requests.sessions.Session,
    *,
    timeout: float,
    request_delay: float,
) -> list[ListingEntry]:
    from openpyxl import load_workbook

    response = get_with_retry(
        session,
        HKEX_SECURITIES_LIST_URL,
        headers=DEFAULT_HEADERS,
        timeout=max(timeout, 60),
        request_delay=request_delay,
    )
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    for _ in range(2):
        next(rows)
    headers = [str(value or "").strip() for value in next(rows)]
    index = {header: pos for pos, header in enumerate(headers)}
    listings: list[ListingEntry] = []
    for row in rows:
        code = str(row[index["Stock Code"]] or "").strip()
        category = str(row[index["Category"]] or "").strip()
        sub_category = str(row[index["Sub-Category"]] or "").strip()
        if not code or category != "Equity" or "Equity Securities" not in sub_category:
            continue
        code = code.zfill(5)
        listings.append(
            ListingEntry(
                code=code,
                secid=f"116.{code}",
                market="HK",
                official_name=row[index["Name of Securities"]],
                board=sub_category,
                trading_currency=row[index["Trading Currency"]],
                listing_source="HKEX",
            )
        )
    return listings


def chunked(items: list[ListingEntry], size: int) -> list[list[ListingEntry]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def fetch_a_share_selector_market_snapshots(
    *,
    markets: list[str],
    min_market_cap: float,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float,
) -> tuple[list[MarketSnapshot], list[dict[str, Any]]]:
    requested = {market for market in markets if market in {"SH", "SZ"}}
    if not requested:
        return [], []

    snapshots: list[MarketSnapshot] = []
    errors: list[dict[str, Any]] = []
    page_number = 1
    page_size = DEFAULT_PAGE_SIZE
    while True:
        params = {
            "type": "RPTA_PCNEW_STOCKSELECT",
            "sty": A_SHARE_SELECTOR_FIELDS,
            "filter": f"(TOTAL_MARKET_CAP>={int(min_market_cap)})",
            "source": "SELECT_SECURITIES",
            "client": "WEB",
            "hyversion": "new",
            "st": "TOTAL_MARKET_CAP",
            "sr": -1,
            "p": page_number,
            "ps": page_size,
        }
        try:
            payload = get_json(session, EASTMONEY_XUANGU_LIST_URL, params=params, request_delay=request_delay, timeout=timeout)
        except Exception as exc:
            errors.append({"market": ",".join(sorted(requested)), "phase": "market_snapshot_xuangu", "error": str(exc)})
            break

        result = payload.get("result") or {}
        rows = result.get("data") or []
        for row in rows:
            try:
                snapshot = parse_a_share_selector_snapshot(row)
            except Exception as exc:
                errors.append({"market": "A", "phase": "market_snapshot_xuangu_parse", "error": str(exc)})
                continue
            if snapshot is None or snapshot.market not in requested:
                continue
            if snapshot.market_cap is None or snapshot.market_cap < min_market_cap:
                continue
            snapshots.append(snapshot)

        if not rows or result.get("nextpage") is False:
            break
        count = to_float(result.get("count"))
        if count is not None and page_number >= math.ceil(count / page_size):
            break
        page_number += 1

    snapshots.sort(key=lambda item: item.market_cap or 0, reverse=True)
    return snapshots, errors


def fetch_hk_tencent_market_snapshots(
    listings: list[ListingEntry],
    *,
    min_market_cap: float,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float,
) -> tuple[list[MarketSnapshot], list[dict[str, Any]]]:
    listings_by_symbol = {f"hk{listing.code.lower()}": listing for listing in listings if listing.market == "HK"}
    snapshots: list[MarketSnapshot] = []
    errors: list[dict[str, Any]] = []
    symbols = list(listings_by_symbol)
    headers = {**DEFAULT_HEADERS, "Referer": "https://gu.qq.com/", "Accept": "*/*"}

    for index in range(0, len(symbols), DEFAULT_PAGE_SIZE):
        batch = symbols[index : index + DEFAULT_PAGE_SIZE]
        if not batch:
            continue
        try:
            response = session.get(f"{TENCENT_QUOTE_URL}{','.join(batch)}", headers=headers, timeout=timeout)
            response.raise_for_status()
            records = parse_tencent_quote_records(response.text)
        except Exception as exc:
            errors.append({"market": "HK", "phase": "market_snapshot_tencent_quote", "error": str(exc)})
            sleep_if_needed(request_delay)
            continue

        for symbol, fields in records.items():
            try:
                snapshot = parse_tencent_hk_snapshot(fields, listings_by_symbol.get(symbol))
            except Exception as exc:
                errors.append({"market": "HK", "phase": "market_snapshot_tencent_parse", "error": str(exc)})
                continue
            if snapshot is None:
                continue
            if snapshot.market_cap is None or snapshot.market_cap < min_market_cap:
                continue
            snapshots.append(snapshot)
        sleep_if_needed(request_delay)

    snapshots.sort(key=lambda item: item.market_cap or 0, reverse=True)
    return snapshots, errors


def fetch_listing_universe_market_snapshots(
    *,
    markets: list[str],
    min_market_cap: float,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float,
) -> tuple[list[MarketSnapshot], list[dict[str, Any]]]:
    snapshots: list[MarketSnapshot] = []
    errors: list[dict[str, Any]] = []
    a_markets = [market for market in markets if market in {"SH", "SZ"}]
    if a_markets:
        a_snapshots, a_errors = fetch_a_share_selector_market_snapshots(
            markets=a_markets,
            min_market_cap=min_market_cap,
            session=session,
            timeout=timeout,
            request_delay=request_delay,
        )
        snapshots.extend(a_snapshots)
        errors.extend(a_errors)

    if "HK" in markets:
        try:
            hk_listings = fetch_hkex_listings(session, timeout=timeout, request_delay=request_delay)
            hk_snapshots, hk_errors = fetch_hk_tencent_market_snapshots(
                hk_listings,
                min_market_cap=min_market_cap,
                session=session,
                timeout=timeout,
                request_delay=request_delay,
            )
            snapshots.extend(hk_snapshots)
            errors.extend(hk_errors)
        except Exception as exc:
            errors.append({"market": "HK", "phase": "listing_universe", "error": str(exc)})

    snapshots.sort(key=lambda item: item.market_cap or 0, reverse=True)
    return snapshots, errors


def limit_snapshots_per_market(
    snapshots: list[MarketSnapshot],
    *,
    markets: list[str],
    max_count: int,
) -> list[MarketSnapshot]:
    if max_count <= 0:
        return snapshots
    limited: list[MarketSnapshot] = []
    for market in markets:
        market_snapshots = [snapshot for snapshot in snapshots if snapshot.market == market]
        limited.extend(market_snapshots[:max_count])
    return limited


def fetch_market_snapshots(
    *,
    markets: list[str],
    max_count: int,
    sort_by: str,
    min_market_cap: float | None = None,
    request_delay: float = 0.0,
    session: requests.sessions.Session | None = None,
    timeout: float = 20,
) -> tuple[list[MarketSnapshot], list[dict[str, Any]]]:
    client = configure_session(session or requests.Session())
    if min_market_cap is not None and sort_by == "market_cap":
        snapshots, errors = fetch_listing_universe_market_snapshots(
            markets=markets,
            min_market_cap=min_market_cap,
            session=client,
            timeout=timeout,
            request_delay=request_delay,
        )
        return limit_snapshots_per_market(snapshots, markets=markets, max_count=max_count), errors

    snapshots: list[MarketSnapshot] = []
    errors: list[dict[str, Any]] = []
    page_size = DEFAULT_PAGE_SIZE if max_count <= 0 else min(max_count, DEFAULT_PAGE_SIZE)
    max_pages = None if max_count <= 0 else max(1, math.ceil(max_count / page_size))

    for market in markets:
        config = MARKET_CONFIGS[market]
        market_snapshots: list[MarketSnapshot] = []
        page_number = 1
        while max_pages is None or page_number <= max_pages:
            params = {
                "pn": page_number,
                "pz": page_size,
                "po": 1,
                "np": 1,
                "ut": EASTMONEY_QUOTE_UT,
                "fltt": 1,
                "invt": 2,
                "fid": SORT_FIELDS.get(sort_by, "f20"),
                "fs": config["fs"],
                "fields": EASTMONEY_LIST_FIELDS,
                "dect": 1,
                "wbp2u": EASTMONEY_WBP2U,
                "_": int(datetime.now(timezone.utc).timestamp() * 1000),
            }
            try:
                response = client.get(EASTMONEY_LIST_URL, params=params, timeout=timeout)
                response.raise_for_status()
                payload = parse_json_or_jsonp(response.text)
                rows = ((payload.get("data") or {}).get("diff") or [])
            except Exception as exc:
                errors.append({"market": market, "phase": "market_snapshot", "error": f"{EASTMONEY_LIST_URL}: {exc}"})
                break

            reached_market_cap_floor = False
            for row in rows:
                try:
                    snapshot = parse_market_snapshot(row, market)
                except Exception as exc:
                    errors.append({"market": market, "phase": "market_snapshot_parse", "error": str(exc)})
                    continue
                if (
                    sort_by == "market_cap"
                    and min_market_cap is not None
                    and snapshot.market_cap is not None
                    and snapshot.market_cap < min_market_cap
                ):
                    reached_market_cap_floor = True
                    continue
                market_snapshots.append(snapshot)
            if len(rows) < page_size:
                break
            if reached_market_cap_floor:
                break
            if max_count > 0 and len(market_snapshots) >= max_count:
                break
            page_number += 1
            sleep_if_needed(request_delay)
        market_snapshots.sort(key=lambda item: item.market_cap or 0, reverse=True)
        snapshots.extend(market_snapshots if max_count <= 0 else market_snapshots[:max_count])

    return snapshots, errors


def annotate_market_cap_universe(snapshots: list[MarketSnapshot]) -> None:
    by_market: dict[str, list[MarketSnapshot]] = {}
    for snapshot in snapshots:
        by_market.setdefault(snapshot.market, []).append(snapshot)

    for market_snapshots in by_market.values():
        ordered = sorted(market_snapshots, key=lambda item: item.market_cap or 0, reverse=True)
        total = len(ordered)
        if total == 0:
            continue
        for index, snapshot in enumerate(ordered, start=1):
            snapshot.market_cap_rank = index
            snapshot.market_cap_percentile = round(100 * index / total, 4)
            snapshot.selected_for_financial_analysis = True


def get_json(
    session: requests.sessions.Session,
    url: str,
    *,
    params: dict[str, Any] | None = None,
    data: dict[str, Any] | None = None,
    request_delay: float = 0.0,
    timeout: float = 20,
) -> dict[str, Any]:
    try:
        if data is None:
            response = session.get(url, params=params, timeout=timeout)
        else:
            response = session.post(url, data=data, timeout=timeout)
        response.raise_for_status()
        try:
            return parse_json_or_jsonp(response.content.decode("utf-8"))
        except UnicodeDecodeError:
            return parse_json_or_jsonp(response.text)
    finally:
        sleep_if_needed(request_delay)


def fetch_a_company_type(
    snapshot: MarketSnapshot,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float,
) -> str:
    params = {
        "filter": f'(SECUCODE="{snapshot.code}.{snapshot.market}")',
        "client": "APP",
        "source": "HSF10",
        "type": "RPT_F10_PUBLIC_COMPANYTPYE",
        "sty": "ALL",
    }
    payload = get_json(session, EASTMONEY_DATA_URL, params=params, request_delay=request_delay, timeout=timeout)
    return str(payload["result"]["data"][0]["COMPANY_TYPE"])


def fetch_a_statement_rows(
    snapshot: MarketSnapshot,
    *,
    statement_type: str,
    style: str,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float,
) -> list[dict[str, Any]]:
    params = {
        "filter": f'(SECUCODE="{snapshot.code}.{snapshot.market}")(REPORT_TYPE="年报")',
        "client": "APP",
        "source": "HSF10",
        "type": statement_type,
        "sty": style,
        "ps": MIN_FINANCIAL_YEARS + 1,
        "sr": -1,
        "st": "REPORT_DATE",
    }
    payload = get_json(session, EASTMONEY_DATA_URL, params=params, request_delay=request_delay, timeout=timeout)
    return (payload.get("result") or {}).get("data") or []


def fetch_a_financial_rows(
    snapshot: MarketSnapshot,
    *,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float = 0.0,
) -> list[dict[str, Any]]:
    company_type = fetch_a_company_type(snapshot, session, timeout, request_delay)
    statement_suffixes = {"1": "S", "2": "I", "3": "B", "4": "G"}
    if company_type not in statement_suffixes:
        raise RuntimeError(f"Unsupported A-share company type for {snapshot.ticker}: {company_type}")
    statement_suffix = statement_suffixes[company_type]

    income_rows = fetch_a_statement_rows(
        snapshot,
        statement_type="RPT_F10_FINANCE_MAINFINADATA",
        style="APP_F10_MAINFINADATA",
        session=session,
        timeout=timeout,
        request_delay=request_delay,
    )
    cash_rows = fetch_a_statement_rows(
        snapshot,
        statement_type=f"RPT_F10_FINANCE_{statement_suffix}CASHFLOW",
        style=f"APP_F10_{statement_suffix}CASHFLOW",
        session=session,
        timeout=timeout,
        request_delay=request_delay,
    )
    balance_rows = fetch_a_statement_rows(
        snapshot,
        statement_type=f"RPT_F10_FINANCE_{statement_suffix}BALANCE",
        style=f"F10_FINANCE_{statement_suffix}BALANCE",
        session=session,
        timeout=timeout,
        request_delay=request_delay,
    )

    by_year: dict[str, dict[str, Any]] = {}
    for row in income_rows:
        year = str(row.get("REPORT_DATE") or "")[:4]
        if not year:
            continue
        by_year.setdefault(year, {"year": year, "currency": "CNY"})
        by_year[year].update(
            {
                "net_profit": to_float(row.get("PARENTNETPROFIT")),
                "roe": to_float(row.get("ROEJQ")),
                "roic": to_float(row.get("ROIC")),
                "debt_to_assets_percent": to_float(row.get("ZCFZL")),
                "payout_ratio": None,
            }
        )
    for row in cash_rows:
        year = str(row.get("REPORT_DATE") or "")[:4]
        if not year:
            continue
        operating_cash_flow = to_float(row.get("NETCASH_OPERATE"))
        capex = to_float(row.get("CONSTRUCT_LONG_ASSET"))
        by_year.setdefault(year, {"year": year, "currency": "CNY"})
        by_year[year].update(
            {
                "operating_cash_flow": operating_cash_flow,
                "capex": capex,
                "free_cash_flow": operating_cash_flow - capex
                if operating_cash_flow is not None and capex is not None
                else None,
            }
        )
    for row in balance_rows:
        year = str(row.get("REPORT_DATE") or "")[:4]
        if not year:
            continue
        total_assets = to_float(row.get("TOTAL_ASSETS"))
        total_liabilities = to_float(row.get("TOTAL_LIABILITIES"))
        goodwill = to_float(row.get("GOODWILL")) or 0.0
        equity = total_assets - total_liabilities if total_assets is not None and total_liabilities is not None else None
        by_year.setdefault(year, {"year": year, "currency": "CNY"})
        by_year[year].update(
            {
                "total_assets": total_assets,
                "total_liabilities": total_liabilities,
                "goodwill": goodwill,
                "equity": equity,
            }
        )
    return [by_year[year] for year in sorted(by_year.keys(), reverse=True)]


def fetch_hk_company_type(
    snapshot: MarketSnapshot,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float,
) -> str:
    payload = get_json(
        session,
        f"{EASTMONEY_HK_DATA_URL}/GetCompanyType",
        data={"fc": snapshot.code.zfill(5), "color": "w"},
        request_delay=request_delay,
        timeout=timeout,
    )
    return str(payload["Result"]["CompanyType"])


def fetch_hk_financial_rows(
    snapshot: MarketSnapshot,
    *,
    session: requests.sessions.Session,
    timeout: float,
    request_delay: float = 0.0,
) -> list[dict[str, Any]]:
    corp_type = fetch_hk_company_type(snapshot, session, timeout, request_delay)
    if corp_type != "4":
        return []

    post_data = {
        "fc": snapshot.code.zfill(5),
        "color": "w",
        "corpType": corp_type,
        "reportDateType": 6,
        "endDate": "",
        "latestCount": MIN_FINANCIAL_YEARS + 1,
        "reportTimeTypeCode": "",
    }
    main_payload = get_json(
        session,
        f"{EASTMONEY_HK_DATA_URL}/GetZhuYaoZhiBiaoList",
        data=post_data,
        request_delay=request_delay,
        timeout=timeout,
    )
    cash_payload = get_json(
        session,
        f"{EASTMONEY_HK_DATA_URL}/GetXianJinLiuLiangList",
        data=post_data,
        request_delay=request_delay,
        timeout=timeout,
    )
    balance_payload = get_json(
        session,
        f"{EASTMONEY_HK_DATA_URL}/GetZiChanFuZhaiList",
        data=post_data,
        request_delay=request_delay,
        timeout=timeout,
    )

    by_year: dict[str, dict[str, Any]] = {}
    for row in ((main_payload.get("Result") or {}).get("ZhuYaoZhiBiaoList_QiYe") or []):
        year = str(row.get("Reportdate") or "")[:4]
        if not year:
            continue
        currency = normalize_currency(row.get("Currency"), snapshot.currency)
        roe = to_float(row.get("Roe"))
        roic = to_float(row.get("Roic"))
        debt_to_assets = to_float(row.get("Debttoassets"))
        if debt_to_assets is not None and abs(debt_to_assets) < 1:
            debt_to_assets *= 100
        by_year.setdefault(year, {"year": year, "currency": currency})
        by_year[year].update(
            {
                "net_profit": to_float(row.get("Parentnetprofit")),
                "roe": roe * 100 if roe is not None and abs(roe) < 1 else roe,
                "roic": roic * 100 if roic is not None and abs(roic) < 1 else roic,
                "debt_to_assets_percent": debt_to_assets,
                "payout_ratio": to_float(row.get("Diviratio")),
            }
        )
    for row in ((cash_payload.get("Result") or {}).get("DataList") or []):
        year = str(row.get("REPORTDATE") or "")[:4]
        if not year:
            continue
        operating_cash_flow = to_float(row.get("CS003999") or row.get("CS002999"))
        buy_capex = sum(to_float(row.get(key)) or 0 for key in ("CS005005", "CS005007"))
        sell_capex = sum(to_float(row.get(key)) or 0 for key in ("CS005004", "CS005006"))
        capex = buy_capex - sell_capex
        by_year.setdefault(year, {"year": year, "currency": normalize_currency(row.get("CURRENCY"), snapshot.currency)})
        by_year[year].update(
            {
                "operating_cash_flow": operating_cash_flow,
                "capex": capex,
                "free_cash_flow": operating_cash_flow - capex if operating_cash_flow is not None else None,
            }
        )
    for row in ((balance_payload.get("Result") or {}).get("DataList") or []):
        year = str(row.get("REPORTDATE") or "")[:4]
        if not year:
            continue
        total_assets = to_float(row.get("BS004009999"))
        total_liabilities = to_float(row.get("BS004025999"))
        goodwill = to_float(row.get("BS004001005")) or 0.0
        equity = total_assets - total_liabilities if total_assets is not None and total_liabilities is not None else None
        by_year.setdefault(year, {"year": year, "currency": normalize_currency(row.get("CURRENCY"), snapshot.currency)})
        by_year[year].update(
            {
                "total_assets": total_assets,
                "total_liabilities": total_liabilities,
                "goodwill": goodwill,
                "equity": equity,
            }
        )
    return [by_year[year] for year in sorted(by_year.keys(), reverse=True)]


def fetch_financial_rows(
    snapshot: MarketSnapshot,
    *,
    session: requests.sessions.Session | None = None,
    request_delay: float = 0.0,
    timeout: float = 20,
) -> list[dict[str, Any]]:
    client = configure_session(session or requests.Session())
    if snapshot.market in {"SH", "SZ"}:
        return fetch_a_financial_rows(snapshot, session=client, request_delay=request_delay, timeout=timeout)
    if snapshot.market == "HK":
        return fetch_hk_financial_rows(snapshot, session=client, request_delay=request_delay, timeout=timeout)
    return []
