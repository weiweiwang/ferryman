from __future__ import annotations

import importlib.util
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "screen_stock_candidates.py"


def load_screen_module():
    spec = importlib.util.spec_from_file_location("screen_stock_candidates_test", SCRIPT_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load stock candidate screener from {SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def sample_snapshot(module, **overrides):
    values = {
        "ticker": "000001.SZ",
        "code": "000001",
        "name": "Quality Co",
        "market": "SZ",
        "currency": "CNY",
        "price": 10.0,
        "pe": 10.0,
        "pb": 2.0,
        "market_cap": 1_000_000_000.0,
        "float_market_cap": 900_000_000.0,
        "industry": "软件服务",
    }
    values.update(overrides)
    return module.MarketSnapshot(**values)


def sample_financial_rows():
    return [
        {
            "year": "2025",
            "currency": "CNY",
            "net_profit": 80_000_000.0,
            "operating_cash_flow": 95_000_000.0,
            "free_cash_flow": 75_000_000.0,
            "roe": 18.0,
            "roic": 15.0,
            "total_assets": 500_000_000.0,
            "total_liabilities": 150_000_000.0,
            "goodwill": 10_000_000.0,
            "equity": 350_000_000.0,
            "payout_ratio": 0.3,
        },
        {
            "year": "2024",
            "currency": "CNY",
            "net_profit": 82_000_000.0,
            "operating_cash_flow": 94_000_000.0,
            "free_cash_flow": 76_000_000.0,
            "roe": 19.0,
            "roic": 15.5,
            "total_assets": 480_000_000.0,
            "total_liabilities": 150_000_000.0,
            "goodwill": 10_000_000.0,
            "equity": 330_000_000.0,
        },
        {
            "year": "2023",
            "currency": "CNY",
            "net_profit": 78_000_000.0,
            "operating_cash_flow": 90_000_000.0,
            "free_cash_flow": 73_000_000.0,
            "roe": 17.0,
            "roic": 14.0,
            "total_assets": 460_000_000.0,
            "total_liabilities": 140_000_000.0,
            "goodwill": 8_000_000.0,
            "equity": 320_000_000.0,
        },
        {
            "year": "2022",
            "currency": "CNY",
            "net_profit": 79_000_000.0,
            "operating_cash_flow": 93_000_000.0,
            "free_cash_flow": 74_000_000.0,
            "roe": 18.0,
            "roic": 14.5,
            "total_assets": 440_000_000.0,
            "total_liabilities": 130_000_000.0,
            "goodwill": 8_000_000.0,
            "equity": 310_000_000.0,
        },
        {
            "year": "2021",
            "currency": "CNY",
            "net_profit": 81_000_000.0,
            "operating_cash_flow": 92_000_000.0,
            "free_cash_flow": 75_000_000.0,
            "roe": 20.0,
            "roic": 16.0,
            "total_assets": 420_000_000.0,
            "total_liabilities": 125_000_000.0,
            "goodwill": 8_000_000.0,
            "equity": 295_000_000.0,
        },
    ]


def risk_free_payload():
    return {"ok": True, "multipleCaps": {"conservativeN2": 15.58}}


def walk_keys(value):
    if isinstance(value, dict):
        for key, item in value.items():
            yield key
            yield from walk_keys(item)
    elif isinstance(value, list):
        for item in value:
            yield from walk_keys(item)


class FakeListResponse:
    def __init__(self, payload):
        self.payload = payload
        self.text = json.dumps(payload)
        self.content = self.text.encode("utf-8")

    def raise_for_status(self):
        return None

    def json(self):
        return self.payload


class RecordingFailListSession:
    def __init__(self):
        self.headers = {}
        self.calls = []

    def get(self, url, params, timeout):
        self.calls.append((url, params, timeout))
        raise RuntimeError("provider request failed")


class AlwaysFailListSession:
    def __init__(self):
        self.headers = {}

    def get(self, url, params, timeout):
        raise RuntimeError("requests host failed")


def test_screener_does_not_import_django_or_radar_runtime():
    script_dir = SCRIPT_PATH.parent
    source = "\n".join(
        (script_dir / filename).read_text(encoding="utf-8")
        for filename in (
            "screen_stock_candidates.py",
            "screen_stock_common.py",
            "screen_stock_metrics.py",
            "screen_stock_parsers.py",
            "screen_stock_providers.py",
        )
    )

    for forbidden in (
        "django",
        "celery",
        "mysql",
        "radar.models",
        "insights.",
        "subprocess",
        "TENCENT_QUOTE_URL",
        "HKEX_SECURITIES_LIST_URL",
        "EASTMONEY_XUANGU_LIST_URL",
        "EASTMONEY_LIST_URL",
        "https://push2.eastmoney.com/api/qt/clist/get",
        "https://data.eastmoney.com/dataapi/xuangu/list",
        "https://qt.gtimg.cn",
        "https://www.hkex.com.hk",
        "fetch_hkex_listings",
        "fetch_hk_tencent_market_snapshots",
        "fetch_a_share_selector_market_snapshots",
    ):
        assert forbidden not in source
    assert "--" + "enrich-limit" not in source
    assert "curl_get_json" not in source
    assert "EASTMONEY_LIST_URLS" not in source


def test_market_snapshot_fetcher_uses_eastmoney_webguest_host():
    module = load_screen_module()

    class PrimaryListSession:
        def __init__(self):
            self.headers = {}
            self.calls = []

        def get(self, url, params, timeout):
            self.calls.append((url, params, timeout))
            return FakeListResponse({"data": {"diff": [{"f12": "601398", "f14": "工商银行", "f20": 1000}]}})

    session = PrimaryListSession()

    snapshots, errors = module.fetch_market_snapshots(
        markets=["SH"],
        max_count=1,
        sort_by="market_cap",
        session=session,
    )

    assert errors == []
    assert len(session.calls) == 1
    url, params, _ = session.calls[0]
    assert url == module.EASTMONEY_WEBGUEST_LIST_URL
    assert params["ut"] == module.EASTMONEY_QUOTE_UT
    assert params["fltt"] == 1
    assert params["dect"] == 1
    assert params["wbp2u"] == module.EASTMONEY_WBP2U
    assert params["timil"] == 1
    assert params["cb"] == "jQuery1124"
    assert snapshots[0].ticker == "601398.SH"


def test_market_snapshot_fetcher_uses_one_webguest_endpoint_for_all_markets():
    module = load_screen_module()

    class AllMarketListSession:
        def __init__(self):
            self.headers = {}
            self.calls = []

        def get(self, url, params, timeout):
            self.calls.append((url, params, timeout))
            fs = params["fs"]
            if "m:1+" in fs:
                rows = [{"f12": "601398", "f13": 1, "f14": "工商银行", "f20": 1000, "f100": "银行Ⅱ"}]
            elif "m:0+" in fs:
                rows = [{"f12": "300750", "f13": 0, "f14": "宁德时代", "f20": 900, "f100": "电池"}]
            elif "m:116" in fs:
                rows = [{"f12": "00700", "f13": 116, "f14": "腾讯控股", "f20": 800, "f100": "软件服务"}]
            elif "m:105" in fs:
                rows = [{"f12": "NVDA", "f13": 105, "f14": "英伟达", "f20": 700, "f100": "信息技术"}]
            else:
                raise AssertionError(f"unexpected fs {fs}")
            return FakeListResponse({"data": {"diff": rows}})

    session = AllMarketListSession()

    snapshots, errors = module.fetch_market_snapshots(
        markets=["SH", "SZ", "HK", "US"],
        max_count=1,
        sort_by="market_cap",
        min_market_cap=1,
        session=session,
    )

    assert errors == []
    assert len(session.calls) == 4
    assert {call[0] for call in session.calls} == {module.MARKET_SNAPSHOT_ENDPOINT}
    assert all(call[1]["timil"] == 1 for call in session.calls)
    assert all(call[1]["cb"] == "jQuery1124" for call in session.calls)
    assert {snapshot.market for snapshot in snapshots} == {"SH", "SZ", "HK", "US"}
    assert {snapshot.source for snapshot in snapshots} == {"eastmoney"}
    assert {snapshot.industry for snapshot in snapshots} == {"银行Ⅱ", "电池", "软件服务", "信息技术"}


def test_market_snapshot_fetcher_uses_webguest_host_for_us_market_cap_floor():
    module = load_screen_module()

    class UsListSession:
        def __init__(self):
            self.headers = {}
            self.calls = []

        def get(self, url, params, timeout):
            self.calls.append((url, params, timeout))
            return FakeListResponse(
                {
                    "data": {
                        "diff": [
                            {
                                "f1": 3,
                                "f2": 194830,
                                "f12": "NVDA",
                                "f13": 105,
                                "f14": "英伟达",
                                "f20": 4_714_886_000_000,
                                "f23": 2412,
                                "f100": "信息技术",
                                "f115": 2954,
                                "f152": 2,
                            }
                        ]
                    }
                }
            )

    session = UsListSession()

    snapshots, errors = module.fetch_market_snapshots(
        markets=["US"],
        max_count=1,
        sort_by="market_cap",
        min_market_cap=1,
        session=session,
    )

    assert errors == []
    assert len(session.calls) == 1
    url, params, _ = session.calls[0]
    assert url == module.EASTMONEY_WEBGUEST_LIST_URL
    assert params["timil"] == 1
    assert params["cb"] == "jQuery1124"
    assert snapshots[0].ticker == "NVDA.O"
    assert snapshots[0].currency == "USD"
    assert snapshots[0].industry == "信息技术"


def test_market_snapshot_fetcher_uses_webguest_host_for_hk_industry_list():
    module = load_screen_module()

    class HkListSession:
        def __init__(self):
            self.headers = {}
            self.calls = []

        def get(self, url, params, timeout):
            self.calls.append((url, params, timeout))
            return FakeListResponse(
                {
                    "data": {
                        "diff": [
                            {
                                "f1": 3,
                                "f2": 431200,
                                "f12": "00700",
                                "f13": 116,
                                "f14": "腾讯控股",
                                "f20": 3_920_571_663_439,
                                "f23": 308,
                                "f100": "软件服务",
                                "f115": 1491,
                                "f152": 2,
                            }
                        ]
                    }
                }
            )

    session = HkListSession()

    snapshots, errors = module.screen_stock_providers.fetch_eastmoney_list_market_snapshots(
        markets=["HK"],
        max_count=1,
        sort_by="market_cap",
        min_market_cap=1,
        session=session,
    )

    assert errors == []
    assert len(session.calls) == 1
    url, params, _ = session.calls[0]
    assert url == module.EASTMONEY_WEBGUEST_LIST_URL
    assert params["timil"] == 1
    assert params["cb"] == "jQuery1124"
    assert snapshots[0].ticker == "00700.HK"
    assert snapshots[0].industry == "软件服务"


def test_market_snapshot_fetcher_does_not_try_alternate_path_when_primary_fails():
    module = load_screen_module()
    session = RecordingFailListSession()

    snapshots, errors = module.fetch_market_snapshots(
        markets=["SH"],
        max_count=1,
        sort_by="market_cap",
        session=session,
    )

    assert snapshots == []
    assert len(session.calls) == 1
    assert session.calls[0][0] == module.EASTMONEY_WEBGUEST_LIST_URL
    assert errors[0]["phase"] == "market_snapshot"
    assert "provider request failed" in errors[0]["error"]


def test_market_snapshot_fetcher_keeps_max_count_per_market():
    module = load_screen_module()

    class MultiMarketSession:
        def __init__(self):
            self.headers = {}

        def get(self, url, params, timeout):
            fs = params["fs"]
            if "m:1+" in fs:
                rows = [
                    {"f12": "600001", "f14": "SH One", "f20": 400},
                    {"f12": "600002", "f14": "SH Two", "f20": 300},
                ]
            else:
                rows = [
                    {"f12": "000001", "f14": "SZ One", "f20": 200},
                    {"f12": "000002", "f14": "SZ Two", "f20": 100},
                ]
            return FakeListResponse({"data": {"diff": rows}})

    snapshots, errors = module.fetch_market_snapshots(
        markets=["SH", "SZ"],
        max_count=2,
        sort_by="market_cap",
        session=MultiMarketSession(),
    )

    assert errors == []
    assert [snapshot.ticker for snapshot in snapshots] == ["600001.SH", "600002.SH", "000001.SZ", "000002.SZ"]


def test_market_snapshot_fetcher_uses_eastmoney_list_when_max_count_is_zero():
    module = load_screen_module()

    class FullUniverseListSession:
        def __init__(self):
            self.headers = {}
            self.calls = []

        def get(self, url, params, timeout):
            self.calls.append((url, params, timeout))
            return FakeListResponse(
                {
                    "data": {
                        "diff": [
                            {"f12": "600001", "f14": "SH One", "f20": 500, "f100": "软件服务"},
                            {"f12": "600002", "f14": "SH Two", "f20": 400, "f100": "软件服务"},
                            {"f12": "600003", "f14": "SH Small", "f20": 200, "f100": "软件服务"},
                        ]
                    }
                }
            )

    session = FullUniverseListSession()
    snapshots, errors = module.fetch_market_snapshots(
        markets=["SH"],
        max_count=0,
        sort_by="market_cap",
        min_market_cap=300,
        session=session,
    )

    assert errors == []
    assert session.calls[0][0] == module.EASTMONEY_WEBGUEST_LIST_URL
    assert [snapshot.ticker for snapshot in snapshots] == ["600001.SH", "600002.SH"]
    assert [snapshot.source for snapshot in snapshots] == ["eastmoney", "eastmoney"]


def test_screen_stocks_returns_stable_failure_when_all_market_snapshots_fail():
    module = load_screen_module()

    payload = module.screen_stocks(
        markets=["HK"],
        max_count=1,
        sort_by="market_cap",
        min_market_cap=100_000_000,
        session=AlwaysFailListSession(),
    )

    assert payload["ok"] is False
    assert payload["phase"] == "market_snapshot"
    assert payload["error"] == "No market snapshots fetched from Eastmoney for requested markets."
    assert payload["summary"]["total"] == 0
    assert payload["results"] == []
    assert payload["data_limits"]["errors"][0]["market"] == "HK"


def test_screen_stocks_returns_failure_when_ulist_universe_returns_no_rows(monkeypatch):
    module = load_screen_module()

    monkeypatch.setattr(
        module,
        "fetch_market_snapshots",
        lambda **kwargs: ([], [{"market": "ALL", "phase": "market_snapshot_ulist", "error": "empty reply"}]),
    )

    payload = module.screen_stocks(
        markets=["SH", "SZ", "HK"],
        max_count=0,
        sort_by="market_cap",
        min_market_cap=100_000_000,
    )

    assert payload["ok"] is False
    assert payload["phase"] == "market_snapshot"
    assert payload["error"] == "No market snapshots fetched; see data_limits.errors."
    assert payload["results"] == []


def test_cli_returns_nonzero_when_market_snapshot_fetch_fails(monkeypatch, capsys):
    module = load_screen_module()

    monkeypatch.setattr(
        module,
        "fetch_market_snapshots",
        lambda **kwargs: ([], [{"market": "HK", "phase": "market_snapshot", "error": "empty reply"}]),
    )

    exit_code = module.main(["--markets", "HK", "--max-count", "1"])
    captured = capsys.readouterr()
    payload = json.loads(captured.out)

    assert exit_code == 1
    assert payload["ok"] is False
    assert payload["phase"] == "market_snapshot"


def test_parse_json_or_jsonp_accepts_passivemoney_callback_shape():
    module = load_screen_module()

    payload = module.parse_json_or_jsonp('jQuery112404724607974107744_1610264100930({"data":{"diff":[]}});')

    assert payload == {"data": {"diff": []}}


def test_parse_market_snapshot_maps_eastmoney_fields_to_snake_case_contract():
    module = load_screen_module()

    snapshot = module.parse_market_snapshot(
        {"f12": "000001", "f14": "Ping An", "f2": 12.3, "f115": 9.8, "f23": 1.1, "f20": 1000, "f21": 800, "f100": "银行"},
        "SZ",
    )

    assert snapshot.ticker == "000001.SZ"
    assert snapshot.currency == "CNY"
    assert snapshot.pe == 9.8
    assert snapshot.pb == 1.1
    assert snapshot.market_cap == 1000


def test_parse_market_snapshot_maps_us_exchange_suffix_and_currency():
    module = load_screen_module()

    snapshot = module.parse_market_snapshot(
        {
            "f12": "AAPL",
            "f13": 105,
            "f14": "苹果",
            "f2": 212.4,
            "f115": 28.5,
            "f23": 45.0,
            "f20": 3_100_000_000_000,
            "f100": "信息技术",
        },
        "US",
    )

    assert snapshot.ticker == "AAPL.O"
    assert snapshot.code == "AAPL"
    assert snapshot.market == "US"
    assert snapshot.currency == "USD"
    assert snapshot.price == 212.4
    assert snapshot.market_cap == 3_100_000_000_000
    assert snapshot.industry == "信息技术"


def test_parse_market_snapshot_scales_official_eastmoney_quote_values():
    module = load_screen_module()

    a_snapshot = module.parse_market_snapshot(
        {
            "f1": 2,
            "f2": 705,
            "f12": "601398",
            "f14": "工商银行",
            "f20": 2_512_664_112_477,
            "f23": 65,
            "f115": 677,
            "f152": 2,
        },
        "SH",
    )
    hk_snapshot = module.parse_market_snapshot(
        {
            "f1": 3,
            "f2": 430200,
            "f12": "00700",
            "f14": "腾讯控股",
            "f20": 3_911_479_428_598,
            "f23": 307,
            "f115": 1488,
            "f152": 2,
        },
        "HK",
    )

    assert a_snapshot.price == 7.05
    assert a_snapshot.pe == 6.77
    assert a_snapshot.pb == 0.65
    assert hk_snapshot.price == 430.2
    assert hk_snapshot.pe == 14.88
    assert hk_snapshot.pb == 3.07


def test_parse_market_snapshot_detects_hk_secondary_quote_currencies():
    module = load_screen_module()

    rmb_counter = module.parse_market_snapshot({"f12": "80700", "f14": "腾讯控股-R"}, "HK")
    usd_counter = module.parse_market_snapshot({"f12": "90700", "f14": "腾讯控股-U"}, "HK")

    assert rmb_counter.ticker == "80700.HK"
    assert rmb_counter.currency == "CNY"
    assert usd_counter.currency == "USD"


def test_parses_eastmoney_hk_currency_and_chinese_amount_units():
    module = load_screen_module()

    assert module.normalize_currency("人民币", "HKD") == "CNY"
    assert module.normalize_currency("港元", "CNY") == "HKD"
    assert module.to_float("2248.42亿") == 224_842_000_000.0
    assert module.to_float("39.13%") == 39.13


def test_build_result_item_marks_candidate_without_quality_or_valuation_flags():
    module = load_screen_module()

    item = module.build_result_item(
        sample_snapshot(module),
        financial_rows=sample_financial_rows(),
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "CANDIDATE"
    assert "quality_flags" not in item
    assert "valuation_flags" not in item
    assert "quality_score" not in item
    assert "valuation_score" not in item
    assert "screen_score" not in item
    assert item["metrics"]["pe"] == 10.0
    assert item["metrics"]["expected_return"] is not None


@pytest.mark.parametrize("market,currency", [("SH", "CNY"), ("HK", "HKD"), ("US", "USD")])
def test_build_result_item_marks_missing_industry_as_data_gap_for_any_market(market, currency):
    module = load_screen_module()
    rows = [row | {"currency": currency} for row in sample_financial_rows()]

    item = module.build_result_item(
        sample_snapshot(module, market=market, currency=currency, industry=None),
        financial_rows=rows,
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "CANDIDATE"
    assert "industry_unavailable" in item["data_gaps"]
    assert item["reject_reasons"] == []


def test_build_result_item_does_not_reject_non_positive_pe_by_itself():
    module = load_screen_module()

    item = module.build_result_item(
        sample_snapshot(module, pe=-3.0),
        financial_rows=sample_financial_rows(),
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "CANDIDATE"
    assert item["reject_reasons"] == []
    assert item["metrics"]["pe"] == -3.0
    assert item["metrics"]["expected_return"] is None


def test_build_result_item_keeps_high_debt_as_metric_not_reject_reason():
    module = load_screen_module()
    rows = [row | {"total_assets": 500_000_000.0, "total_liabilities": 450_000_000.0} for row in sample_financial_rows()]

    item = module.build_result_item(
        sample_snapshot(module),
        financial_rows=rows,
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "CANDIDATE"
    assert item["metrics"]["debt_to_assets"] == 0.9
    assert "high_debt" not in item["reject_reasons"]


def test_build_result_item_keeps_three_negative_fcf_years_as_metric_not_reject_reason():
    module = load_screen_module()
    rows = sample_financial_rows()
    rows[0] = rows[0] | {"free_cash_flow": -1.0}
    rows[1] = rows[1] | {"free_cash_flow": -1.0}
    rows[2] = rows[2] | {"free_cash_flow": -1.0}

    item = module.build_result_item(
        sample_snapshot(module),
        financial_rows=rows,
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "CANDIDATE"
    assert item["metrics"]["negative_fcf_years"] == 3
    assert "too_many_negative_fcf_years" not in item["reject_reasons"]


def test_build_result_item_keeps_low_fcf_to_profit_as_metric_not_reject_reason():
    module = load_screen_module()
    rows = [row | {"free_cash_flow": 10_000_000.0} for row in sample_financial_rows()]

    item = module.build_result_item(
        sample_snapshot(module),
        financial_rows=rows,
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "CANDIDATE"
    assert item["metrics"]["fcf_to_profit"] < 0.25
    assert "weak_fcf_conversion" not in item["reject_reasons"]


def test_a_share_financial_rows_do_not_fallback_when_company_type_is_unknown(monkeypatch):
    module = load_screen_module()
    snapshot = sample_snapshot(module, market="SH", ticker="600001.SH", code="600001")

    monkeypatch.setattr(module.screen_stock_providers, "fetch_a_company_type", lambda *args, **kwargs: "9")

    with pytest.raises(RuntimeError, match="Unsupported A-share company type"):
        module.fetch_a_financial_rows(
            snapshot,
            session=module.configure_session(module.requests.Session()),
            timeout=1,
            request_delay=0,
        )


def test_a_share_financial_rows_propagate_company_type_errors(monkeypatch):
    module = load_screen_module()
    snapshot = sample_snapshot(module, market="SH", ticker="600001.SH", code="600001")

    def fail_company_type(*args, **kwargs):
        raise RuntimeError("company type unavailable")

    monkeypatch.setattr(module.screen_stock_providers, "fetch_a_company_type", fail_company_type)

    with pytest.raises(RuntimeError, match="company type unavailable"):
        module.fetch_a_financial_rows(
            snapshot,
            session=module.configure_session(module.requests.Session()),
            timeout=1,
            request_delay=0,
        )


def test_us_financial_rows_use_pc_f10_statements_to_build_five_year_fcf():
    module = load_screen_module()
    snapshot = sample_snapshot(module, ticker="AAPL.O", code="AAPL", market="US", currency="USD")

    years = ["2025", "2024", "2023", "2022", "2021"]

    class FakeUsF10Session:
        def __init__(self):
            self.headers = {}
            self.calls = []

        def get(self, url, params, timeout):
            self.calls.append((url, params, timeout))
            report_name = params["reportName"]
            filter_clause = params["filter"]
            if report_name == "RPT_USSK_FN_CASHFLOW" and "REPORT in" not in filter_clause:
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                {
                                    "SECUCODE": "AAPL.O",
                                    "REPORT": f"{year}/FY",
                                    "REPORT_DATE": f"{year}-09-30",
                                    "CURRENCY": "美元",
                                    "DATE_TYPE_CODE": "001",
                                }
                                for year in years
                            ]
                        }
                    }
                )
            if report_name == "RPT_USF10_FN_INCOME":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                item
                                for year in years
                                for item in (
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004001999",
                                        "AMOUNT": 390_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004015999",
                                        "AMOUNT": 100_000_000_000.0,
                                    },
                                )
                            ]
                        }
                    }
                )
            if report_name == "RPT_USSK_FN_CASHFLOW":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                item
                                for year in years
                                for item in (
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "003999",
                                        "AMOUNT": 120_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "005002",
                                        "AMOUNT": -12_000_000_000.0,
                                    },
                                )
                            ]
                        }
                    }
                )
            if report_name == "RPT_USF10_FN_BALANCE":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                item
                                for year in years
                                for item in (
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004001001",
                                        "AMOUNT": 35_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004005999",
                                        "AMOUNT": 350_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004007007",
                                        "AMOUNT": 10_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004009005",
                                        "AMOUNT": 90_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004011999",
                                        "AMOUNT": 250_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004017999",
                                        "AMOUNT": 100_000_000_000.0,
                                    },
                                )
                            ]
                        }
                    }
                )
            if report_name == "RPT_USF10_FN_GMAININDICATOR":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                {
                                    "SECUCODE": "AAPL.O",
                                    "REPORT_DATE": f"{year}-09-30",
                                    "DATE_TYPE_CODE": "001",
                                    "ROE_AVG": 95.0,
                                    "DEBT_ASSET_RATIO": 71.4,
                                }
                                for year in years
                            ]
                        }
                    }
                )
            raise AssertionError(f"unexpected reportName {report_name}")

    rows = module.fetch_us_financial_rows(snapshot, session=FakeUsF10Session(), timeout=1, request_delay=0)

    assert [row["year"] for row in rows] == years
    assert rows[0]["currency"] == "USD"
    assert rows[0]["net_profit"] == 100_000_000_000.0
    assert rows[0]["operating_cash_flow"] == 120_000_000_000.0
    assert rows[0]["capex"] == 12_000_000_000.0
    assert rows[0]["free_cash_flow"] == 108_000_000_000.0
    assert rows[0]["cash_and_equivalents"] == 35_000_000_000.0
    assert rows[0]["total_debt"] == 100_000_000_000.0
    assert rows[0]["roe"] == 95.0


def test_us_financial_rows_keep_extra_report_when_latest_year_is_incomplete():
    module = load_screen_module()
    snapshot = sample_snapshot(module, ticker="AAPL.O", code="AAPL", market="US", currency="USD")

    report_years = ["2026", "2025", "2024", "2023", "2022", "2021"]
    complete_years = report_years[1:]

    class FakeIncompleteLatestUsSession:
        def __init__(self):
            self.headers = {}

        def get(self, url, params, timeout):
            report_name = params["reportName"]
            filter_clause = params["filter"]
            if report_name == "RPT_USSK_FN_CASHFLOW" and "REPORT in" not in filter_clause:
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                {
                                    "SECUCODE": "AAPL.O",
                                    "REPORT": f"{year}/FY",
                                    "REPORT_DATE": f"{year}-09-30",
                                    "CURRENCY": "美元",
                                    "DATE_TYPE_CODE": "001",
                                }
                                for year in report_years
                            ]
                        }
                    }
                )
            if report_name == "RPT_USF10_FN_INCOME":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                item
                                for year in report_years
                                for item in (
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004001999",
                                        "AMOUNT": 390_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004015999",
                                        "AMOUNT": 100_000_000_000.0,
                                    },
                                )
                            ]
                        }
                    }
                )
            if report_name == "RPT_USSK_FN_CASHFLOW":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                item
                                for year in complete_years
                                for item in (
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "003999",
                                        "AMOUNT": 120_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "005002",
                                        "AMOUNT": -12_000_000_000.0,
                                    },
                                )
                            ]
                        }
                    }
                )
            if report_name == "RPT_USF10_FN_BALANCE":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                item
                                for year in report_years
                                for item in (
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004005999",
                                        "AMOUNT": 350_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004011999",
                                        "AMOUNT": 250_000_000_000.0,
                                    },
                                    {
                                        "REPORT": f"{year}/FY",
                                        "REPORT_DATE": f"{year}-09-30",
                                        "STD_ITEM_CODE": "004017999",
                                        "AMOUNT": 100_000_000_000.0,
                                    },
                                )
                            ]
                        }
                    }
                )
            if report_name == "RPT_USF10_FN_GMAININDICATOR":
                return FakeListResponse(
                    {
                        "result": {
                            "data": [
                                {
                                    "REPORT_DATE": f"{year}-09-30",
                                    "DATE_TYPE_CODE": "001",
                                    "ROE_AVG": 95.0,
                                }
                                for year in report_years
                            ]
                        }
                    }
                )
            raise AssertionError(f"unexpected reportName {report_name}")

    rows = module.fetch_us_financial_rows(
        snapshot,
        session=FakeIncompleteLatestUsSession(),
        timeout=1,
        request_delay=0,
    )
    item = module.build_result_item(
        snapshot,
        financial_rows=rows,
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert [row["year"] for row in rows] == report_years
    assert rows[0]["free_cash_flow"] is None
    assert item["status"] == "CANDIDATE"
    assert item["metrics"]["complete_financial_years"] == 5
    assert "missing_5y_financial_rows" not in item["data_gaps"]


def test_build_result_item_uses_reject_reasons_for_obvious_rejects():
    module = load_screen_module()

    item = module.build_result_item(
        sample_snapshot(module, name="*ST Bad Co", price=0, pe=-3),
        financial_rows=[],
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "REJECTED"
    assert set(item["reject_reasons"]) >= {"no_valid_price", "st_or_special_treatment"}
    assert "non_positive_pe" not in item["reject_reasons"]
    assert "quality_flags" not in item
    assert "valuation_flags" not in item


def test_build_result_item_uses_data_gaps_for_missing_financial_history():
    module = load_screen_module()

    item = module.build_result_item(
        sample_snapshot(module),
        financial_rows=sample_financial_rows()[:3],
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "INSUFFICIENT_DATA"
    assert "missing_5y_financial_rows" in item["data_gaps"]
    assert item["reject_reasons"] == []


def test_build_result_item_uses_recent_five_complete_years_when_latest_year_is_incomplete():
    module = load_screen_module()
    rows = sample_financial_rows()
    incomplete_latest = rows[0] | {"free_cash_flow": None}
    older_complete = rows[-1] | {
        "year": "2020",
        "net_profit": 83_000_000.0,
        "operating_cash_flow": 96_000_000.0,
        "free_cash_flow": 77_000_000.0,
    }

    item = module.build_result_item(
        sample_snapshot(module),
        financial_rows=[incomplete_latest, *rows[1:], older_complete],
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "CANDIDATE"
    assert item["metrics"]["complete_financial_years"] == 5
    assert item["metrics"]["avg_fcf_5y"] == 75_000_000.0
    assert "missing_5y_financial_rows" not in item["data_gaps"]


def test_industry_review_required_takes_precedence_over_generic_financial_rules():
    module = load_screen_module()
    weak_bank_rows = [row | {"free_cash_flow": -1.0} for row in sample_financial_rows()]

    item = module.build_result_item(
        sample_snapshot(module, industry="银行"),
        financial_rows=weak_bank_rows,
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    assert item["status"] == "INDUSTRY_REVIEW_REQUIRED"
    assert "industry_review_required" in item["data_gaps"]
    assert item["reject_reasons"] == []


def test_result_payload_uses_snake_case_without_legacy_camel_case():
    module = load_screen_module()

    item = module.build_result_item(
        sample_snapshot(module),
        financial_rows=sample_financial_rows(),
        risk_free=risk_free_payload(),
        min_market_cap=100_000_000,
    )

    forbidden = {
        "metricSnapshot",
        "peTtm",
        "roeMean",
        "roeStd",
        "qualityFlags",
        "valuationFlags",
        "rejectReasons",
        "nextResearchChecks",
        "quality_flags",
        "valuation_flags",
        "quality_score",
        "valuation_score",
        "screen_score",
    }
    assert forbidden.isdisjoint(set(walk_keys(item)))


def test_screen_stocks_analyzes_all_market_cap_floor_rows(monkeypatch):
    module = load_screen_module()
    snapshots = [
        sample_snapshot(module, ticker=f"00000{index}.SZ", code=f"00000{index}", market_cap=market_cap)
        for index, market_cap in enumerate([500, 400, 300, 200, 100], start=1)
    ]
    fetched = []

    monkeypatch.setattr(module, "fetch_market_snapshots", lambda **kwargs: (snapshots, []))

    def fake_fetch_financial_rows(snapshot, **kwargs):
        fetched.append(snapshot.ticker)
        return sample_financial_rows()

    monkeypatch.setattr(module, "fetch_financial_rows", fake_fetch_financial_rows)
    monkeypatch.setattr(module, "fetch_risk_free_rate", lambda *args, **kwargs: risk_free_payload() | {"ok": True})

    payload = module.screen_stocks(
        markets=["SZ"],
        max_count=5,
        sort_by="market_cap",
        min_market_cap=1,
    )

    assert fetched == ["000001.SZ", "000002.SZ", "000003.SZ", "000004.SZ", "000005.SZ"]
    assert payload["data_limits"]["market_snapshot_provider"] == module.MARKET_SNAPSHOT_PROVIDER
    assert payload["data_limits"]["market_snapshot_endpoint"] == module.MARKET_SNAPSHOT_ENDPOINT
    assert payload["data_limits"]["financial_analysis_scope"] == "all_rows_at_or_above_market_cap_floor"
    assert payload["data_limits"]["financial_analysis_selected_count"] == 5
    assert payload["data_limits"]["analyzed_count"] == 5
    assert payload["data_limits"]["financial_fetch_attempts"] == 5
    assert payload["results"][0]["market_cap_rank"] == 1
    assert payload["results"][0]["market_cap_percentile"] == 20.0
    assert payload["results"][0]["analyzed"] is True
    assert all("outside_top_20_percent_by_market_cap" not in item["data_gaps"] for item in payload["results"])


def test_screen_stocks_passes_request_delay_to_financial_fetch(monkeypatch):
    module = load_screen_module()
    snapshot = sample_snapshot(module)
    delays = []

    monkeypatch.setattr(module, "fetch_market_snapshots", lambda **kwargs: ([snapshot], []))

    def fake_fetch_financial_rows(snapshot, **kwargs):
        delays.append(kwargs["request_delay"])
        return sample_financial_rows()

    monkeypatch.setattr(module, "fetch_financial_rows", fake_fetch_financial_rows)
    monkeypatch.setattr(module, "fetch_risk_free_rate", lambda *args, **kwargs: risk_free_payload() | {"ok": True})

    payload = module.screen_stocks(
        markets=["SZ"],
        max_count=1,
        sort_by="market_cap",
        min_market_cap=1,
        request_delay=0.25,
    )

    assert payload["ok"] is True
    assert delays == [0.25]
    assert payload["data_limits"]["request_delay_seconds"] == 0.25


def test_cli_writes_json_and_xlsx_outputs(monkeypatch, tmp_path, capsys):
    module = load_screen_module()
    snapshot = sample_snapshot(module)

    monkeypatch.setattr(module, "fetch_market_snapshots", lambda **kwargs: ([snapshot], []))
    monkeypatch.setattr(module, "fetch_financial_rows", lambda *args, **kwargs: sample_financial_rows())
    monkeypatch.setattr(module, "fetch_risk_free_rate", lambda *args, **kwargs: risk_free_payload() | {"ok": True})

    json_out = tmp_path / "screen.json"
    xlsx_out = tmp_path / "screen.xlsx"
    exit_code = module.main(
        [
            "--markets",
            "SH",
            "SZ",
            "HK",
            "--max-count",
            "1",
            "--min-market-cap",
            "100000000",
            "--json-out",
            str(json_out),
            "--xlsx-out",
            str(xlsx_out),
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    assert json.loads(captured.out)["summary"]["candidate"] == 1
    payload = json.loads(json_out.read_text(encoding="utf-8"))
    assert payload["results"][0]["ticker"] == "000001.SZ"
    workbook = load_workbook(xlsx_out)
    assert "All" in workbook.sheetnames
    assert workbook["Candidate"].max_row == 2
    headers = [cell.value for cell in workbook["Candidate"][1]]
    assert "avg_fcf_5y" in headers
    assert "quality_flags" not in headers
    assert "valuation_flags" not in headers
    assert "screen_score" not in headers


def test_cli_checkpoint_run_writes_jsonl_and_default_xlsx(monkeypatch, tmp_path, capsys):
    module = load_screen_module()
    snapshot = sample_snapshot(module)

    monkeypatch.setattr(module, "fetch_market_snapshots", lambda **kwargs: ([snapshot], []))
    monkeypatch.setattr(module, "fetch_financial_rows", lambda *args, **kwargs: sample_financial_rows())
    monkeypatch.setattr(module, "fetch_risk_free_rate", lambda *args, **kwargs: risk_free_payload() | {"ok": True})

    run_dir = tmp_path / "reports" / "stock-screen" / "2026-07-02"
    exit_code = module.main(
        [
            "--markets",
            "SZ",
            "--max-count",
            "1",
            "--run-date",
            "2026-07-02",
            "--run-dir",
            str(run_dir),
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    assert json.loads(captured.out)["data_limits"]["run_dir"] == str(run_dir)
    assert (run_dir / "universe.json").exists()
    assert (run_dir / "enrich.jsonl").exists()
    assert not (run_dir / "progress.json").exists()
    assert (run_dir / "stock-screen-2026-07-02.xlsx").exists()
    rows = [json.loads(line) for line in (run_dir / "enrich.jsonl").read_text(encoding="utf-8").splitlines()]
    assert rows[0]["ticker"] == "000001.SZ"
    assert rows[0]["enrich_status"] == "ok"


def test_cli_logs_progress_to_stderr_without_polluting_stdout(monkeypatch, capsys):
    module = load_screen_module()
    snapshots = [
        sample_snapshot(module, ticker="000001.SZ", code="000001", name="One", market_cap=300),
        sample_snapshot(module, ticker="000002.SZ", code="000002", name="Two", market_cap=200),
        sample_snapshot(module, ticker="000003.SZ", code="000003", name="Three", market_cap=100),
    ]

    monkeypatch.setattr(module, "fetch_market_snapshots", lambda **kwargs: (snapshots, []))

    def fake_fetch_financial_rows(snapshot, **kwargs):
        if snapshot.ticker == "000002.SZ":
            raise RuntimeError("timeout")
        return sample_financial_rows()

    monkeypatch.setattr(module, "fetch_financial_rows", fake_fetch_financial_rows)
    monkeypatch.setattr(module, "fetch_risk_free_rate", lambda *args, **kwargs: risk_free_payload() | {"ok": True})

    exit_code = module.main(
        [
            "--markets",
            "SZ",
            "--max-count",
            "3",
            "--min-market-cap",
            "1",
            "--progress-interval",
            "2",
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    payload = json.loads(captured.out)
    assert payload["data_limits"]["progress_interval"] == 2
    assert payload["summary"]["total"] == 3
    assert "[stock-screen] universe live total=3 markets=SZ" in captured.err
    assert "enrich 1/3" not in captured.err
    assert "enrich 2/3 ok=1 failed=1" in captured.err
    assert "event=failed" in captured.err
    assert "enrich 3/3 ok=2 failed=1" in captured.err
    assert "[stock-screen] complete total=3 ok=2 failed=1" in captured.err


def test_fresh_run_dir_resets_old_checkpoint(monkeypatch, tmp_path):
    module = load_screen_module()
    snapshot = sample_snapshot(module)
    run_dir = tmp_path / "reports" / "stock-screen" / "2026-07-02"
    run_dir.mkdir(parents=True)
    (run_dir / "universe.json").write_text('{"snapshots":[{"ticker":"OLD"}]}\n', encoding="utf-8")
    (run_dir / "enrich.jsonl").write_text('{"ticker":"OLD","enrich_status":"ok","result":{}}\n', encoding="utf-8")
    (run_dir / "progress.json").write_text('{"completed":1}\n', encoding="utf-8")

    monkeypatch.setattr(module, "fetch_market_snapshots", lambda **kwargs: ([snapshot], []))
    monkeypatch.setattr(module, "fetch_financial_rows", lambda *args, **kwargs: sample_financial_rows())
    monkeypatch.setattr(module, "fetch_risk_free_rate", lambda *args, **kwargs: risk_free_payload() | {"ok": True})

    payload = module.screen_stocks(
        markets=["SZ"],
        max_count=1,
        sort_by="market_cap",
        min_market_cap=1,
        run_date="2026-07-02",
        run_dir=run_dir,
        resume=False,
    )

    assert payload["ok"] is True
    universe = json.loads((run_dir / "universe.json").read_text(encoding="utf-8"))
    assert universe["market_snapshot_provider"] == module.MARKET_SNAPSHOT_PROVIDER
    assert universe["market_snapshot_endpoint"] == module.MARKET_SNAPSHOT_ENDPOINT
    assert universe["snapshots"][0]["ticker"] == "000001.SZ"
    rows = [json.loads(line) for line in (run_dir / "enrich.jsonl").read_text(encoding="utf-8").splitlines()]
    assert [row["ticker"] for row in rows] == ["000001.SZ"]
    assert not (run_dir / "progress.json").exists()


def test_resume_uses_universe_checkpoint_skips_ok_and_retries_failed(monkeypatch, tmp_path):
    module = load_screen_module()
    first = sample_snapshot(module, ticker="000001.SZ", code="000001", name="First")
    second = sample_snapshot(module, ticker="000002.SZ", code="000002", name="Second")
    module.annotate_market_cap_universe([first, second])
    run_dir = tmp_path / "reports" / "stock-screen" / "2026-07-02"
    module.write_universe(
        run_dir,
        run_date="2026-07-02",
        markets=["SZ"],
        snapshots=[first, second],
        errors=[],
        min_market_cap=1,
    )
    first_result = module.build_result_item(
        first,
        financial_rows=sample_financial_rows(),
        risk_free=risk_free_payload(),
        min_market_cap=1,
    )
    first_result["status"] = "REJECTED"
    first_result["quality_score"] = 70
    first_result["valuation_score"] = 0
    first_result["screen_score"] = 42
    first_result["quality_flags"] = ["stable_roe"]
    first_result["valuation_flags"] = []
    first_result["reject_reasons"] = ["not_enough_quality_or_valuation_signals"]
    module.checkpoint_result(
        run_dir,
        run_date="2026-07-02",
        snapshot=first,
        enrich_status="ok",
        result=first_result,
        attempt=1,
    )
    failed_result = module.financial_fetch_failed_item(second, min_market_cap=1, error="timeout")
    module.checkpoint_result(
        run_dir,
        run_date="2026-07-02",
        snapshot=second,
        enrich_status="failed",
        result=failed_result,
        attempt=1,
        error="timeout",
    )
    fetched = []

    def fail_if_market_snapshots_are_fetched(**kwargs):
        raise AssertionError("resume should load universe.json instead of fetching snapshots")

    def fake_fetch_financial_rows(snapshot, **kwargs):
        fetched.append(snapshot.ticker)
        return sample_financial_rows()

    monkeypatch.setattr(module, "fetch_market_snapshots", fail_if_market_snapshots_are_fetched)
    monkeypatch.setattr(module, "fetch_financial_rows", fake_fetch_financial_rows)
    monkeypatch.setattr(module, "fetch_risk_free_rate", lambda *args, **kwargs: risk_free_payload() | {"ok": True})

    payload = module.screen_stocks(
        markets=["SZ"],
        max_count=0,
        sort_by="market_cap",
        min_market_cap=1,
        run_date="2026-07-02",
        run_dir=run_dir,
        resume=True,
    )

    assert payload["ok"] is True
    assert fetched == ["000002.SZ"]
    assert {item["ticker"] for item in payload["results"]} == {"000001.SZ", "000002.SZ"}
    refreshed_first = next(item for item in payload["results"] if item["ticker"] == "000001.SZ")
    assert refreshed_first["status"] == "CANDIDATE"
    assert refreshed_first["reject_reasons"] == []
    assert "quality_flags" not in refreshed_first
    assert "valuation_flags" not in refreshed_first
    assert "screen_score" not in refreshed_first
    latest_rows = module.latest_checkpoint_rows(run_dir)
    assert latest_rows["000001.SZ"]["enrich_status"] == "ok"
    assert latest_rows["000002.SZ"]["enrich_status"] == "ok"
    assert latest_rows["000002.SZ"]["attempt"] == 2
    assert not (run_dir / "progress.json").exists()


def test_resume_rejects_cross_day_checkpoint(tmp_path):
    module = load_screen_module()
    snapshot = sample_snapshot(module)
    run_dir = tmp_path / "reports" / "stock-screen" / "2026-07-02"
    module.write_universe(
        run_dir,
        run_date="2026-07-02",
        markets=["SZ"],
        snapshots=[snapshot],
        errors=[],
        min_market_cap=1,
    )

    payload = module.screen_stocks(
        markets=["SZ"],
        max_count=0,
        sort_by="market_cap",
        min_market_cap=1,
        run_date="2026-07-03",
        run_dir=run_dir,
        resume=True,
    )

    assert payload["ok"] is False
    assert payload["phase"] == "resume_checkpoint"
    assert "Cannot resume run_date 2026-07-03" in payload["error"]
    assert payload["results"] == []


def test_resume_rejects_checkpoint_from_missing_market_snapshot_provider(tmp_path):
    module = load_screen_module()
    snapshot = sample_snapshot(module)
    run_dir = tmp_path / "reports" / "stock-screen" / "2026-07-02"
    module.write_universe(
        run_dir,
        run_date="2026-07-02",
        markets=["SZ"],
        snapshots=[snapshot],
        errors=[],
        min_market_cap=1,
    )
    universe_path = run_dir / "universe.json"
    universe = json.loads(universe_path.read_text(encoding="utf-8"))
    universe.pop("market_snapshot_provider")
    universe.pop("market_snapshot_endpoint")
    universe_path.write_text(json.dumps(universe), encoding="utf-8")

    payload = module.screen_stocks(
        markets=["SZ"],
        max_count=0,
        sort_by="market_cap",
        min_market_cap=1,
        run_date="2026-07-02",
        run_dir=run_dir,
        resume=True,
    )

    assert payload["ok"] is False
    assert payload["phase"] == "resume_checkpoint"
    assert "different market snapshot provider" in payload["error"]
    assert payload["data_limits"]["market_snapshot_provider"] == module.MARKET_SNAPSHOT_PROVIDER
    assert payload["data_limits"]["market_snapshot_endpoint"] == module.MARKET_SNAPSHOT_ENDPOINT
    assert payload["data_limits"]["checkpoint_market_snapshot_provider"] is None
    assert payload["data_limits"]["checkpoint_market_snapshot_endpoint"] is None
    assert payload["results"] == []


def test_financial_cache_reuses_rows_without_refetch(monkeypatch, tmp_path):
    module = load_screen_module()
    snapshot = sample_snapshot(module)
    cache_dir = tmp_path / "reports" / "stock-screen" / "cache" / "financials"
    module.write_financial_rows_cache(cache_dir, snapshot, sample_financial_rows())

    def fail_fetch(*args, **kwargs):
        raise AssertionError("cached financial rows should avoid live fetch")

    monkeypatch.setattr(module, "fetch_financial_rows", fail_fetch)

    rows, from_cache = module.fetch_financial_rows_cached(
        snapshot,
        cache_dir=cache_dir,
        cache_max_age_days=7,
        session=module.configure_session(module.requests.Session()),
        request_delay=0,
        timeout=1,
    )

    assert from_cache is True
    assert rows == sample_financial_rows()


def test_financial_cache_refetches_when_fetched_at_is_stale(monkeypatch, tmp_path):
    module = load_screen_module()
    snapshot = sample_snapshot(module)
    cache_dir = tmp_path / "reports" / "stock-screen" / "cache" / "financials"
    module.write_financial_rows_cache(cache_dir, snapshot, sample_financial_rows())
    cache_path = module.financial_cache_path(cache_dir, snapshot)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    payload["fetched_at"] = (datetime.now(timezone.utc) - timedelta(days=8)).isoformat()
    cache_path.write_text(json.dumps(payload), encoding="utf-8")
    fetched = []

    def fake_fetch_financial_rows(snapshot, **kwargs):
        fetched.append(snapshot.ticker)
        return [row | {"year": "2026"} for row in sample_financial_rows()]

    monkeypatch.setattr(module, "fetch_financial_rows", fake_fetch_financial_rows)

    rows, from_cache = module.fetch_financial_rows_cached(
        snapshot,
        cache_dir=cache_dir,
        cache_max_age_days=7,
        session=module.configure_session(module.requests.Session()),
        request_delay=0,
        timeout=1,
    )

    assert from_cache is False
    assert fetched == ["000001.SZ"]
    assert {row["year"] for row in rows} == {"2026"}


def test_financial_cache_ttl_can_be_disabled(monkeypatch, tmp_path):
    module = load_screen_module()
    snapshot = sample_snapshot(module)
    cache_dir = tmp_path / "reports" / "stock-screen" / "cache" / "financials"
    module.write_financial_rows_cache(cache_dir, snapshot, sample_financial_rows())
    cache_path = module.financial_cache_path(cache_dir, snapshot)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    payload["fetched_at"] = (datetime.now(timezone.utc) - timedelta(days=365)).isoformat()
    cache_path.write_text(json.dumps(payload), encoding="utf-8")

    def fail_fetch(*args, **kwargs):
        raise AssertionError("disabled TTL should allow stale cache reuse")

    monkeypatch.setattr(module, "fetch_financial_rows", fail_fetch)

    rows, from_cache = module.fetch_financial_rows_cached(
        snapshot,
        cache_dir=cache_dir,
        cache_max_age_days=0,
        session=module.configure_session(module.requests.Session()),
        request_delay=0,
        timeout=1,
    )

    assert from_cache is True
    assert rows == sample_financial_rows()
